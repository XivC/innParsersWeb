import openpyxl as pyx
from dbf_light import Dbf
def isEmpty(string):
        if string.strip() in ["-", "", " ", "None"]:
            return True
        return False

column_name_relations = {
    "authCapital": "Уставной капитал",
        "phones": "Номера телефонов",
        "address": "Адрес",
        "workers": "Число сотрудников",
        "shortName": "Сокращ. имя",
        "fullName": "Полное имя",
        "ogrn": "ОГРН",
        "status": "Статус",
        "regDate": "Дата регистрации",
        "directors": "Управляющие",
        "holders": "Владельцы",
        "reportYear": "Год отчётности",
        "earnings": "Выручка",
        "govContracts": "Гос. контракты",
        "govBuyings": "Гос. закупки",
        "inn": "ИНН",
        "region": "Регион",
        "city": "Город"
        }
def parseAutoColumns(path, fields):
    fields = fields
    auto_fields = [f for f in fields.keys() if fields[f] == "auto"]
    wb = pyx.open(path)
    ws = wb.active
    col = 1
    while len(auto_fields) > 0:
        if isEmpty(str(ws.cell(row = 1, column = col).value)):
            target = auto_fields.pop(0)
            ws.cell(row = 1, column = col).value = column_name_relations.get(target)
            fields[target] = pyx.utils.cell.get_column_letter(col)
        col += 1
    wb.save(filename = path)
    return fields



def scan(inpath, outpath, fields = {
    "authCapital": "auto",
        "phones": "auto",
        "address": "auto",
        "workers": "auto",
        "shortName": "auto",
        "fullName": "auto",
        "inn": "auto",
        "ogrn": "auto",
        "status": "auto",
        "regDate": "auto",
        "directors": "auto",
        "holders": "auto",
        "reportYear": "auto",
        "earnings": "auto",
        "govContracts": "auto",
        "govBuyings": "auto",
        "region": "auto",
        "city": "auto",
        "inn_search_columns": [],
        "inn_search_key_column": "A",
        


}, services = {
    "inn_parser": True,
    "information_scanner": True,
    "okved_scanner": True,
}):
    print(fields)
    print(services)
    fields = parseAutoColumns(inpath, fields)
    
    
    if services["inn_parser"] == True:
        from inn.Scanner import Scanner
        scanner = Scanner()
        scanner.scan_xlsx_to_xlsx(inpath, fields["inn_search_key_column"], fields["inn_search_columns"], fields["inn"], inpath, ignore = ['1'], number_correction=[fields["phones"]])
    if services["information_scanner"] == True:
        from InfScanner.InformationCollector import InformationScanner
        scanner = InformationScanner()
        scanner.scan(inpath, inpath, fields["inn"], output_file_fields=fields)
    if services["okved_scanner"] == True:
        from InfScanner.InformationCollector import InformationScanner
        print("scanning okveds")
        okved_fields = {
             "main" : "auto",
        "additional": "auto",
        "all_start": "auto"
        }
        okved_fields = parseAutoColumns(inpath, okved_fields)

        scanner = InformationScanner()
        scanner.scanOkveds(inpath, inpath, fields["ogrn"], fields["ogrn"], output_file_fields = okved_fields )
    wb = pyx.open(inpath)
    wb.save(filename = outpath)

if __name__ == "__main__":
    import json
    data = {}
    with Dbf.open('zipbase.dbf') as dbf:
         for rec in dbf:
                    
                data[rec[0]] = (rec[4], rec[7])
    with open("zipbase.json", "w") as f:
        f.write(json.dumps(data))
                    

    '''
    scan("Domains_20.xlsx", "Domains_20_res.xlsx", fields =  {
        "authCapital": "auto",
            "phones": "R",
            "address": "P",
            "workers": "auto",
            "shortName": "auto",
            "fullName": "auto",
            "inn": "H",
            "ogrn": "N",
            "status": "auto",
            "regDate": "auto",
            "directors": "auto",
            "holders": "auto",
            "reportYear": "auto",
            "earnings": "auto",
            "govContracts": "auto",
            "govBuyings": "auto",
            "region": "auto",
            "city": "O",
            "inn_search_columns": ["A", 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'],
            "inn_search_key_column": "A",


    } )
    '''
