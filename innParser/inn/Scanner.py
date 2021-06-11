import requests
from bs4 import BeautifulSoup as bs
import traceback
import sys
import pandas
import json
import random
import time
import threading
import openpyxl as pyx
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import pathlib
class Scanner:
    
    spark = True
    seldon = True
    SBIS = True
    


    def __init__(self):
        #download_path = pathlib.Path(download).absolute()
        
        options = webdriver.ChromeOptions()
        profile_driver_path = pathlib.Path().absolute()
        profile_path = profile_driver_path.joinpath("Profile")
        print("PAAAAATH",profile_path)
        webdriver_path = profile_driver_path.joinpath("chromedriver.exe")
        print(webdriver_path)
        options.add_argument("user-data-dir={}".format(profile_path))     
        self.driver = webdriver.Chrome(executable_path=webdriver_path, chrome_options=options)
        self.sbis_session = requests.Session()
        self.kontur_session = requests.Session()

        self.driver.get("https://online.saby.ru/page/company")
        for cookie in self.driver.get_cookies():
            self.sbis_session.cookies.set(cookie['name'], cookie['value'])
        self.driver.get("https://focus.kontur.ru/")
        for cookie in self.driver.get_cookies():
            self.kontur_session.cookies.set(cookie['name'], cookie['value'])
        self.driver.close()
        

    def scan_xlsx_to_xlsx(self, input_file, key_column, query_columns, column_to_write, output_file, ignore = [], inJson = False, outJson = False, number_correction = False):
        data = Utils.getFromExcel(input_file, key_column, query_columns, ignore, number_correction = number_correction)
        if inJson != False:
            with open(inJson, "w") as f:
                f.write(json.dumps(data))
        result = {}
        self.scan_base(data, result)
        if outJson != False:
            with open(outJson, "w") as f:
                f.write(json.dumps(result))
        Utils.writeExcel(input_file, result, key_column, column_to_write, outpath = output_file)
        


    def scan_base(self, data, result):
        res_spark = {}
        res_seldon = {}
        res_sbis = {}
        res_kontur = {}
        sparkT = threading.Thread(target = self._scanBaseBySystem, args=(data, "spark", res_spark))
        seldonT = threading.Thread(target = self._scanBaseBySystem, args=(data, "seldon", res_seldon))
        sbisT = threading.Thread(target = self._scanBaseBySystem, args=(data, "SBIS", res_sbis))
        konturT = threading.Thread(target = self._scanBaseBySystem, args=(data, "kontur", res_kontur))
        sparkT.start()
        seldonT.start()
        sbisT.start()
        konturT.start()
        while sparkT.is_alive() or seldonT.is_alive() or sbisT.is_alive() or konturT.is_alive():
            continue
        Utils.merge([res_spark, res_seldon, res_sbis, res_kontur], result)
        
    
    def _scanBaseBySystem(self, data, system, result):
        
        for key in list(data.keys()):
            inn = "-"
            for query in data[key]:
                if type(query) != type(""):
                    continue
                if len(query) <= 3:
                    continue
                if system == "spark":
                    inn = self._scanSpark(query)
                elif system == "seldon":
                    inn = self._scanSeldon(query)
                elif system == "SBIS":
                    inn = self._scanSBIS(query)
                elif system == "kontur":
                    inn = self._scanKontur(query)

                if inn != False:
                    break
                else:
                    inn = "-"
            result[key] = inn
            print(key , result[key])
    def search(self, query):
        if len(str(query)) <= 3:
            #print("Плохой запрос")
            return False
        result = False
        if result == False and self.spark:
            result = self._scanSpark(query)
        if result == False and self.seldon:
            result = self._scanSeldon(query)
        if result == False and self.SBIS:
            result = self._scanSBIS(query)
        if result == False:
            result = self._scanKontur(query)
        return result

    def _scanSeldon(self, query):
        #print(query)
    
        
        time.sleep( random.randint(5,10) + random.random())

        from requests.structures import CaseInsensitiveDict
        
        url = "https://basis.myseldon.com/ru/home/searchdata/"

        headers = CaseInsensitiveDict()
        headers["Content-Type"] = "application/x-www-form-urlencoded"

        #data = "pageIndex=1&searchString={query}&pageSize=12&useSnippet=true".format(query = query.replace(" ", "%")).encode("utf-8")
       
        data = {
            'pageIndex': '1',
            'searchString': query,
            'pageSize': '12',
            'useSnippet': 'true'
        }
           
        try:    
            resp = requests.post(url, headers=headers, data=data)
            if resp.status_code == 200:
                try:
                    heap = json.loads(resp.content)['response']
                    if heap['items'] != [] and len(heap['items']) < 4:
                        for company in heap['items']:
                            if company["statusName"] == "Действующая":
                                inn = company["inn"].replace("<b>", "").replace("</b>", "")
                                print("found: seldon")
                                return company["inn"].replace("<b>", "").replace("</b>", "")
                            return False
                    else:
                        
                        return False
                except:
                    traceback.print_exc
                    return False
            else:
                print(resp)
        except:
            traceback.print_exc
            return False

    def _scanSpark(self, query):
        r = requests.get("https://www.spark-interfax.ru/search?Query={query}".format(query = query))
        if r.status_code != 200:
            print("Караул")
            return False
        #with open("spark_dumps/dump " + str(query) + ".html", "wb") as f:
        #    f.write(r.text.encode("utf-8"))
        inn = ""
        soup = bs(r.content, features = "html.parser")
        try:
            info = soup.find("li", {"class": "search-result-list__item"})
            for company in info.find_all("div", {"class":"summary"}):
                fields_inn = company.find("div", {"class":"code"}).find_all("span")
                for i in range(len(fields_inn)):
                    if fields_inn[i].text == "ИНН":
                        inn = fields_inn[i+1].text
                        #print(inn)
                        print("found: spark")
                        return inn
        except:
            #traceback.print_exc(file=sys.stdout)
            return False

    def _scanSBIS(self, query):
        time.sleep( random.randint(2,3) + random.random())
        from requests.structures import CaseInsensitiveDict
        url = "https://online.saby.ru/service/"
        cookie_string = "; ".join([str(x)+"="+str(y) for x,y in self.sbis_session.cookies.items()])
        headers = CaseInsensitiveDict()
        headers["accept"] = "application/json, text/javascript, */*; q=0.01"
        headers["accept-encoding"] = "gzip, deflate, br"
        headers["accept-language"] = "ru-RU;q=0.8,en-US;q=0.5,en;q=0.3"
        headers["Content-Type"] = "application/json; charset=UTF-8"
        headers["cookie"] = cookie_string
        headers["dnt"] = "1"
        headers["origin"] = "https://online.saby.ru"
        headers["referer"] = "https://online.saby.ru/page/company/tab/list/main"
        headers["sec-ch-ua"] = 'Chromium";v="88", "Google Chrome";v="88", ";Not A Brand";v="99"'
        headers["sec-ch-ua-mobile"] = "?0"
        headers["sec-fetch-dest"] = "empty"
        headers["sec-fetch-mode"] = "cors"
        headers["sec-fetch-site"] = "same-origin"
        headers["user-agent"] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36"
        headers["x-requested-with"] = "XMLHttpRequest"
        
        data = '{"jsonrpc":"2.0","protocol":6,"method":"Contractor.ListCompany","params":{"Фильтр":{"d":[[],"'+query+'",[null],true,true,[],[1]],"s":[{"t":{"n":"Массив","t":"Строка"},"n":"Category"},{"t":"Строка","n":"Details"},{"t":{"n":"Массив","t":"Строка"},"n":"KindMan"},{"t":"Логическое","n":"MatchedFields"},{"t":"Логическое","n":"Misspelling"},{"t":{"n":"Массив","t":"Строка"},"n":"Region"},{"t":{"n":"Массив","t":"Число целое"},"n":"State"}],"_type":"record","f":0},"Сортировка":{"d":[[false,"Релевантность",true]],"s":[{"t":"Логическое","n":"l"},{"t":"Строка","n":"n"},{"t":"Логическое","n":"o"}],"_type":"recordset","f":0},"Навигация":{"d":[true,40,0],"s":[{"t":"Логическое","n":"ЕстьЕще"},{"t":"Число целое","n":"РазмерСтраницы"},{"t":"Число целое","n":"Страница"}],"_type":"record","f":0},"ДопПоля":[]},"id":1}'
        data = data.encode("utf-8")
        
       

        try:
            req = requests.Request('POST', url, data=data, headers=headers)
            req_prep = req.prepare()
            #print(req_prep.headers)
            resp = self.sbis_session.send(req_prep)#, data = data)#requests.post(url, headers=headers,data=data)
            
            if resp.status_code != 200:
                print(resp.status_code)
                print(resp.text)
                return False
            data = json.loads(str(resp.content, 'utf-8'))
            companies = data['result']['d']
            #print(companies)
            if len(companies) == 0 or len(companies) > 4:
                return False
            try:
                print("found: sbis")
                return data['result']['d'][0][7]
            except:
                traceback.print_exc()
                return False
        except:
            traceback.print_exc()
            return False
    def _scanKontur(self, query):
        time.sleep( random.randint(2,3) + random.random())
        cookie_string = "; ".join([str(x)+"="+str(y) for x,y in self.kontur_session.cookies.items()])
        headers = {
        'sec-ch-ua': '"Chromium";v="88", "Google Chrome";v="88", ";Not A Brand";v="99"',
        'DNT': '1',
        'accept-language': 'en-GB,en;q=0.9,ru-RU;q=0.8,ru;q=0.7,el-GR;q=0.6,el;q=0.5,en-US;q=0.4',
        'sec-ch-ua-mobile': '?0',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded; charset=utf-8',
        'accept': 'application/json',
        'x-reactapp-version': 'f386f65ed685f3d56f5d',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
        'Accept-Encoding': 'gzip, deflate, br',
        'cookie':cookie_string#.encode("utf-8")
        }
        payload = {
            'country': "RU",
            'query': query
        }
    
        try:
            url = "https://focus.kontur.ru/api/search"
            req = requests.Request('GET', url, params = payload, headers=headers)
            
            req_prep = req.prepare()
            #print(req_prep.headers)
            #r = self.kontur_session.send(req_prep)

            r = requests.get("https://focus.kontur.ru/api/search", headers = headers, params = payload)
            #print(r.text)
            data = json.loads(r.text)["data"]
            if(len(data) == 0 or len(data) > 6):
                return False
            else:
                print("found: kontur")
                return data[0]["requisites"][0]["value"]
        except:
            traceback.print_exc
            return False


        

class Utils:
    @staticmethod 
    def numberCorrection(number):
        if number != None:
            return str(number).replace("(", "").replace(")", "").replace("-", "").replace(" ", "")
        else:
            return number
    
    @staticmethod
    def merge(toMerge, res):
        for dct in toMerge:
            for key in dct.keys():
                if res.get(key) == None:
                    res[key] = dct[key]
                elif res.get(key) == "-":
                    if dct[key] != '-' or dct[key] != None:
                        res[key] = dct[key]

    @staticmethod
    def getFromExcel(path, key_col, query_cols, ignore, number_correction = False):
        print(path)
        wb = pyx.load_workbook(filename = path)
        ws = wb.active
        try: 
            key_column = pyx.utils.column_index_from_string(key_col)
        except:
            traceback.print_exc()
            return {}
        query_columns = []
        for q in query_cols:
            try:
                query_columns.append(pyx.utils.column_index_from_string(q))
            except:
                continue
        number_corrections = []
        if number_correction != False:
            for q in query_cols:
                try:
                    number_corrections.append(pyx.utils.column_index_from_string(q))
                except:
                    continue


        
        result = {}
        for row in range(1, len([x for x in ws.rows]) + 1):
            if str(row) in ignore:
                continue
            key = ws.cell(row = row, column = key_column).value
            queries = []
            for column in query_columns:
                q = ws.cell(row = row, column = column).value
                if column in number_corrections:
                    q = Utils.numberCorrection(q)
                queries.append(q)
            result[key] = queries

            
        return result
    def writeExcel(inpath, data, key_col, result_col, outpath = None):
        if outpath == None:
            outpath = inpath
        wb = pyx.load_workbook(filename = inpath)
        print("aasdsd")
        ws = wb.active
        try: 
            key_column = pyx.utils.column_index_from_string(key_col)
            result_column = pyx.utils.column_index_from_string(result_col)
        except:
            traceback.print_exc()
            return 
        for row in range(1, len([x for x in ws.rows]) + 1):
            for column in range(1, len([x for x in ws.columns]) + 1):
                if column == key_column:
                    if data.get(ws.cell(row = row, column = key_column).value) != None:
                        #print("found!")
                        ws.cell(row = row, column = result_column).value = data[ws.cell(row = row, column = key_column).value]
        wb.save(filename = outpath)



        




if __name__ == "__main__":
    scanner = Scanner()
   
    '''
    seldon = json.loads(open("seldon_result.json").read())
    sbis = json.loads(open("sbis_result.json").read())
    spark = json.loads(open("spark_result.json").read())
    kontur = json.loads(open("kontur_result.json").read())
    
    res = {}
    Utils.merge([seldon, sbis, spark, kontur], res)
    print(res)
    Utils.writeExcel("Domains.xlsx", res, 'A', 'H') 
'''
    '''
    
    import requests
    from requests.structures import CaseInsensitiveDict

    
    wb = pyx.load_workbook(filename = "result_filtered_spark.xlsx")
    ws = wb.active
    flt = dict({})
    for row in range(1, len([x for x in ws.rows])):
        t = ws.cell(row = row, column = 1).value
        inn = ws.cell(row = row, column = 2).value
        
        flt[t] = inn

    j = json.dumps(flt)
    with open("spark_result.json", "w") as f:
        f.write(j)
  '''






