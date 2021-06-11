import openpyxl as pyx
from pathlib import Path
from openpyxl.styles import Alignment

class OkvedXlsxUtils:
    @staticmethod
    def _isEmpty(string):
        if string.strip() in ["-", "", " ", "None"]:
            return True
        return False
    @staticmethod
    def getOkvedsByOgrn(files):

        res = {}
        for file_ in files:
            subres = {
                "main": "",
                "additional": [],
                "all": []
            }
            ogrn = Path(file_).stem
            print(ogrn)
            wb = pyx.load_workbook(filename = file_)
            ws = wb.active
            flag = False
            target_range = []
            for row in ws.rows:
                #print(row[0].coordinate, type(row[0]))
                if row[0].value == "Основной вид деятельности":
                        #print(row[0])
                        ok = str(row[1].value)
                        if (not OkvedXlsxUtils._isEmpty(ok)):
                            subres["main"]= str(row[1].value)
                            #print(ogrn, ok.split()[0])
                            subres["all"].append(ok.split()[0])
                        flag = True

                if row[0].value == "Дополнительные виды деятельности":
                    for rng in ws.merged_cells.ranges:
                        if row[0].coordinate in rng:
                            target_range = rng
                            break
                    ok = str(row[1].value)
                    if (not OkvedXlsxUtils._isEmpty(ok)):
                        subres["additional"].append(ok)
                        #print(ogrn, ok.split()[0])
                        subres["all"].append(ok.split()[0])
            try:
                for c in pyx.worksheet.cell_range.CellRange(str(target_range)).cells:
                    ok = str((ws.cell(row = c[0], column = c[1]+1).value))
                    if (not OkvedXlsxUtils._isEmpty(ok)):
                        subres["additional"].append(ok)
                        print(ogrn, ok.split()[0])
                        subres["all"].append(ok.split()[0])
            except:
                pass
            res[ogrn] = subres
        return res
    
    def toXlsx(data, path, ogrn_field, fields = {
        "main" : "AR",
        "additional": "AS",
        "all_start": "AT"
    }, outpath = None):
        if outpath == None:
            outpath = path
        
        index_fields = {
            "main": ("Основной вид деятельности", pyx.utils.column_index_from_string(fields["main"])),
            "additional": ("Дополнительные виды деятельности", pyx.utils.column_index_from_string(fields["additional"]))
        }
        all_okveds = []
        start_okv = pyx.utils.column_index_from_string(fields["all_start"])
        for x in data.keys():
            all_okveds += data[x]["all"]
        all_okveds = set(all_okveds)
        for okv in all_okveds:
            index_fields[okv] = (okv, start_okv)
            start_okv += 1
        print(index_fields)
        #print(all_okveds)
        wb = pyx.load_workbook(filename = path)
        ws = wb.active
        #wb.save(filename = path+"_okved_backup.xlsx")
        for c in [x for x in ws.columns]:
            for cc in c:
                
                cc.alignment = Alignment(wrapText=True)
        
        for field in index_fields.values():
            print(field)
            ws.cell(row = 1, column = field[1]).value = str(field[0])
        ogrn_column = pyx.utils.column_index_from_string(ogrn_field)
        rows = len([x for x in ws.rows])
        print(rows)
        
        for row in range(2, rows + 1):
            rd = ws.row_dimensions[row] 
            rd.height = 25
            rd.width = 25
            ogrn = str(ws.cell(row = row, column = ogrn_column).value)
            #print(ogrn)
            if ogrn in data.keys():
                print(ogrn, "found")
                ws.cell(row = row, column = index_fields["main"][1]).value = data[ogrn]["main"]
                ws.cell(row = row, column = index_fields["additional"][1]).value = "\n".join(data[ogrn]["additional"])
                for okv in data[ogrn]["all"]:
                    cell = ws.cell(row = row, column = index_fields[okv][1])
                    fill = pyx.styles.fills.PatternFill(patternType='solid', fgColor=pyx.styles.colors.Color(rgb='0000FF00'))
                    cell.fill = fill
                    cell.value = "+"
        wb.save(filename = outpath)







                



def main():
    import glob
    import json
    with open("okved_data.json") as f:
       data = json.loads(f.read())
    #data = OkvedXlsxUtils.getOkvedsByOgrn(glob.glob("downloads/*.xlsx"))
    with open('okved_data.json', "w") as f:
        f.write(json.dumps(data))
    #for x in data.keys():
    #    print(data[x]["all"])
    OkvedXlsxUtils.toXlsx(data, "Domains.xlsx", "N")
if __name__ == "__main__":
    main()