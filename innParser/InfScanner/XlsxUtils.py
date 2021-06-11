import openpyxl as pyx
import json
from openpyxl.styles import Alignment
import traceback
class XlsxUtils:
    @staticmethod
    def _isEmpty(string):
        if string.strip() in ["-", "", " ", "None"]:
            return True
        return False
    @staticmethod
    def XlsxToJson(path, fields = {
        "authCapital": "R",
        "phones": "T",
        "address": "U",
        "workers": "J",
        "shortName": "A",
        "fullName": "B",
        "inn": "C",
        "ogrn": "E",
        "status": "F",
        "regDate": "G",
        "directors": "K",
        "holders": "N",
        "reportYear": "X",
        "earnings": "Z",
        "govContracts": "AH",
        "govBuyings": "AJ"}):
        index_fields = {}
        for key in fields.keys():
            index_fields[key] = pyx.utils.column_index_from_string(fields[key])

        wb = pyx.load_workbook(filename = path)
        ws = wb.active
       
        result = {}
        rows = len([x for x in ws.rows])
        for row in range(2, rows + 1):
            inn = ws.cell(row = row, column = index_fields["inn"]).value
            result[inn] = {
            'phones' : str(ws.cell(row = row, column = index_fields["phones"]).value),
            'ogrn' : str(ws.cell(row = row, column = index_fields["ogrn"]).value),
            'authCapital' : str(ws.cell(row = row, column = index_fields["authCapital"]).value),
            'address' : str(ws.cell(row = row, column = index_fields["address"]).value),
            'workers' : str(ws.cell(row = row, column = index_fields["workers"]).value),
            'shortName' : str(ws.cell(row = row, column = index_fields["shortName"]).value),
            'fullName' : str(ws.cell(row = row, column = index_fields["fullName"]).value),
            'status' : str(ws.cell(row = row, column = index_fields["status"]).value),
            'regDate' : str(ws.cell(row = row, column = index_fields["regDate"]).value).replace("00:00:00", ""),
            'directors' : str(ws.cell(row = row, column = index_fields["directors"]).value),
            'holders' : str(ws.cell(row = row, column = index_fields["holders"]).value).replace("Учредители и участники:",""),
            'reportYear' : str(ws.cell(row = row, column = index_fields["reportYear"]).value),
            'earnings' : str(ws.cell(row = row, column = index_fields["earnings"]).value),
            'govContracts' : str(ws.cell(row = row, column = index_fields["govContracts"]).value),
            'govBuyings' : str(ws.cell(row = row, column = index_fields["govBuyings"]).value),
            }
        js = json.dumps(result)
        with open(path+".json", "w") as f:
            f.write(js)
        return result
        
    def JsonToXlsx(data, path, fields = {
        "authCapital": "AF",
        "phones": "R",
        "address": "P",
        "workers": "AG",
        "shortName": "AH",
        "fullName": "AI",
        "inn": "H",
        "ogrn": "N",
        "status": "AJ",
        "regDate": "AK",
        "directors": "AL",
        "holders": "AM",
        "reportYear": "AN",
        "earnings": "AO",
        "govContracts": "AP",
        "govBuyings": "AQ",
        "city": "AR",
        "region": "AS",}, outpath = None):
        
        if outpath == None:
            outpath = path
        index_fields = {}
        for key in fields.keys():
            try:
                index_fields[key] = pyx.utils.column_index_from_string(fields[key])
            except:
                continue
        try:
            wb = pyx.load_workbook(filename = path)
        except:
            traceback.print_exc()
            wb = pyx.Workbook()
        ws = wb.active
        wb.save(filename = outpath + "_backup.xlsx")
        cells = [c for c in [x for x in ws.columns]]
        
        
        for c in [x for x in ws.columns]:
            for cc in c:
                
                cc.alignment = Alignment(wrapText=True)
        
        #return
        rows = len([x for x in ws.rows])
    
        for row in range(2, rows + 1):
            rd = ws.row_dimensions[row] 
            rd.height = 25
            inn = str(ws.cell(row = row, column = index_fields["inn"]).value)
            if data.get(inn) != None:
                
                for key in index_fields.keys():
                    if key not in data.get(inn).keys():
                    
                        continue
                    cell = ws.cell(row = row, column = index_fields[key])
                  
                    cell_v = str(cell.value)
                    data_v = str(data.get(inn).get(key))
                
                    if not XlsxUtils._isEmpty(data_v):
                        cell.value = data_v
                    #print(cell.value)
        wb.save(filename = outpath)
    @staticmethod
    def ColumnToList(path, column):
        wb = pyx.load_workbook(filename = path)
        ws = wb.active
        result = []
        inn_index = 1
        try:
            inn_index = pyx.utils.column_index_from_string(column)
        except:
            return []
        rows = len([x for x in ws.rows])
        for row in range(2, rows + 1):
            inn = str(ws.cell(row = row, column = inn_index).value)
            if(not XlsxUtils._isEmpty(inn)):
                result.append(inn)
        return result
            
            
            




        
        

        
           
if __name__ == "__main__":
    data = XlsxUtils.XlsxToJson("test_dump.xlsx")
    XlsxUtils.JsonToXlsx(data, "Domains.xlsx")


            
